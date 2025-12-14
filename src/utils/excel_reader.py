"""
Excel/CSV表格读取模块
"""

import os
import csv
from typing import List, Dict, Any
from openpyxl import load_workbook
from src.core.logger import get_logger

logger = get_logger()


class Article:
    """文章数据类"""

    def __init__(self, index: int, title: str, content: str):
        self.index = index
        self.title = title
        self.content = content
        self.published = False
        self.publish_result = None

    def __repr__(self):
        return f"Article({self.index}, {self.title[:20]}...)"


class ExcelReader:
    """Excel/CSV读取器"""

    def __init__(self):
        self.articles: List[Article] = []
        self.file_path: str = ""

    def load(self, file_path: str) -> bool:
        """
        加载Excel或CSV文件

        Args:
            file_path: 文件路径（支持.xlsx, .xls, .csv）

        Returns:
            是否加载成功
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"文件不存在: {file_path}")
                return False

            self.file_path = file_path
            self.articles = []

            # 根据文件扩展名选择加载方式
            ext = os.path.splitext(file_path)[1].lower()

            if ext == '.csv':
                return self._load_csv(file_path)
            elif ext in ['.xlsx', '.xls']:
                return self._load_excel(file_path)
            else:
                logger.error(f"不支持的文件格式: {ext}")
                return False

        except Exception as e:
            logger.error(f"加载文件失败: {e}")
            return False

    def _load_csv(self, file_path: str) -> bool:
        """加载CSV文件（跳过第一行标题）"""
        try:
            # 尝试不同的编码
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']

            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as f:
                        reader = csv.reader(f)
                        # 跳过第一行（标题行）
                        next(reader, None)

                        for idx, row in enumerate(reader, start=1):
                            if len(row) >= 2 and row[0] and row[1]:
                                title = str(row[0]).strip()
                                content = str(row[1]).strip()
                                if title and content:
                                    article = Article(idx, title, content)
                                    self.articles.append(article)

                    logger.info(f"成功加载CSV文件 {len(self.articles)} 篇文章 (编码: {encoding})")
                    return True
                except UnicodeDecodeError:
                    self.articles = []  # 清空重试
                    continue

            logger.error("无法识别CSV文件编码")
            return False

        except Exception as e:
            logger.error(f"加载CSV失败: {e}")
            return False

    def _load_excel(self, file_path: str) -> bool:
        """加载Excel文件（跳过第一行标题）"""
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            # min_row=2 跳过第一行标题
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if row[0] and row[1]:
                    title = str(row[0]).strip()
                    content = str(row[1]).strip()
                    if title and content:
                        article = Article(idx, title, content)
                        self.articles.append(article)

            wb.close()
            logger.info(f"成功加载Excel文件 {len(self.articles)} 篇文章")
            return True

        except Exception as e:
            logger.error(f"加载Excel失败: {e}")
            return False
    
    def get_articles(self) -> List[Article]:
        """获取所有文章"""
        return self.articles
    
    def get_unpublished_articles(self, count: int) -> List[Article]:
        """获取指定数量的未发布文章"""
        unpublished = [a for a in self.articles if not a.published]
        return unpublished[:count]
    
    def mark_as_published(self, article: Article, result: str = "success"):
        """标记文章为已发布"""
        article.published = True
        article.publish_result = result
        logger.info(f"文章已标记为发布: {article.title[:20]}... - {result}")

